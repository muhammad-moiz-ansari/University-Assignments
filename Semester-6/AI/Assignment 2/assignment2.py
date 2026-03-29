import random
from itertools import combinations
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────
# ZONE DATA
# ─────────────────────────────────────────
# Key: zone number (1–8)
# Value: dict with description and capacity in kg

zones = {
    1: {"name": "Heavy Item Floor Storage",           "capacity": 120},
    2: {"name": "Standard Rack Storage",              "capacity": 80},
    3: {"name": "Fragile Item Shelf",                 "capacity": 80},
    4: {"name": "Temperature Controlled Storage",     "capacity": 80},
    5: {"name": "Hazardous Material Storage",         "capacity": 80},
    6: {"name": "Fast-Moving Product Area",           "capacity": 60},
    7: {"name": "Bulk Dry Storage",                   "capacity": 150},
    8: {"name": "Refrigerated Loading Dock",          "capacity": 100},
}

# ─────────────────────────────────────────
# PRODUCT DATA
# ─────────────────────────────────────────
# Index 0 = P1, Index 1 = P2, ..., Index 19 = P20
# Each product has:
#   name, weight, category,
#   fragile, hazardous, temp_controlled, high_demand

products = [
    {"name": "Glass Bottles",     "weight": 20, "category": "Beverage",    "fragile": True,  "hazardous": False, "temp_controlled": False, "high_demand": False},
    {"name": "Frozen Meat",       "weight": 30, "category": "Food",        "fragile": False, "hazardous": False, "temp_controlled": True,  "high_demand": True },
    {"name": "Cleaning Acid",     "weight": 10, "category": "Chemical",    "fragile": False, "hazardous": True,  "temp_controlled": False, "high_demand": False},
    {"name": "Rice Bags",         "weight": 50, "category": "Grocery",     "fragile": False, "hazardous": False, "temp_controlled": False, "high_demand": True },
    {"name": "Ceramic Plates",    "weight": 15, "category": "Kitchenware", "fragile": True,  "hazardous": False, "temp_controlled": False, "high_demand": False},
    {"name": "Ice Cream",         "weight": 25, "category": "Food",        "fragile": False, "hazardous": False, "temp_controlled": True,  "high_demand": True },
    {"name": "Detergent",         "weight": 12, "category": "Chemical",    "fragile": False, "hazardous": True,  "temp_controlled": False, "high_demand": False},
    {"name": "Chips Carton",      "weight":  8, "category": "Snacks",      "fragile": False, "hazardous": False, "temp_controlled": False, "high_demand": True },
    {"name": "Olive Oil Bottles", "weight": 18, "category": "Grocery",     "fragile": True,  "hazardous": False, "temp_controlled": False, "high_demand": False},
    {"name": "Industrial Bleach", "weight": 22, "category": "Chemical",    "fragile": False, "hazardous": True,  "temp_controlled": False, "high_demand": False},
    {"name": "Yogurt Cartons",    "weight": 14, "category": "Food",        "fragile": False, "hazardous": False, "temp_controlled": True,  "high_demand": True },
    {"name": "Flour Bags",        "weight": 45, "category": "Grocery",     "fragile": False, "hazardous": False, "temp_controlled": False, "high_demand": True },
    {"name": "Wine Bottles",      "weight": 16, "category": "Beverage",    "fragile": True,  "hazardous": False, "temp_controlled": False, "high_demand": False},
    {"name": "Paint Cans",        "weight": 28, "category": "Chemical",    "fragile": False, "hazardous": True,  "temp_controlled": False, "high_demand": False},
    {"name": "Biscuit Boxes",     "weight":  6, "category": "Snacks",      "fragile": False, "hazardous": False, "temp_controlled": False, "high_demand": True },
    {"name": "Motor Oil",         "weight": 35, "category": "Chemical",    "fragile": False, "hazardous": True,  "temp_controlled": False, "high_demand": False},
    {"name": "Frozen Fish",       "weight": 20, "category": "Food",        "fragile": False, "hazardous": False, "temp_controlled": True,  "high_demand": True },
    {"name": "Bubble Wrap Rolls", "weight": 10, "category": "Packaging",   "fragile": True,  "hazardous": False, "temp_controlled": False, "high_demand": False},
    {"name": "Wheat Sacks",       "weight": 40, "category": "Grocery",     "fragile": False, "hazardous": False, "temp_controlled": False, "high_demand": True },
    {"name": "Hand Sanitizer",    "weight":  8, "category": "Chemical",    "fragile": True,  "hazardous": False, "temp_controlled": False, "high_demand": False},
]

# ─────────────────────────────────────────
# QUICK REFERENCE SETS (for fitness checks)
# ─────────────────────────────────────────
# These sets store product INDICES (0-based) for fast lookup

fragile_products     = {i for i, p in enumerate(products) if p["fragile"]}
hazardous_products   = {i for i, p in enumerate(products) if p["hazardous"]}
temp_ctrl_products   = {i for i, p in enumerate(products) if p["temp_controlled"]}
high_demand_products = {i for i, p in enumerate(products) if p["high_demand"]}
heavy_products       = {i for i, p in enumerate(products) if p["weight"] > 40}

# Only Frozen Meat (1), Ice Cream (5), Yogurt Cartons (10), Frozen Fish (16) are eligible for Z8
z8_eligible_products = {1, 5, 10, 16}

# Food and Chemical product indices (for incompatibility check)
food_products     = {i for i, p in enumerate(products) if p["category"] == "Food"}
chemical_products = {i for i, p in enumerate(products) if p["category"] == "Chemical"}

num_zones = 8
num_products = len(products)

# ─────────────────────────────────────────
# FUNCTION DEFINITIONS
# ─────────────────────────────────────────

def calculate_fitness(chromo):
    fitness = 0

    # 5.1 - Zone Weight Capacity
    # Weights assigned to each zone (0-indexed)
    weights = [0 for i in range(num_zones + 1)]  # +1 to make it 1-indexed for easier mapping
    for i in range(len(chromo)):
        weights[chromo[i]] += products[i]["weight"]
    
    for i in range(1, num_zones + 1):
        diff = weights[i] - zones[i]["capacity"]
        if diff > 0:
            fitness += diff * 2
    
    # 5.2 - Fragile Product Protection
    for i in range(len(chromo)):
        if chromo[i] != 3 and i in fragile_products:
            fitness += 8
    
    # 5.3 - Hazardous Material Isolation
    for i in range(len(chromo)):
        if chromo[i] != 5 and i in hazardous_products:
            fitness += 10

    # 5.4 - Temperature-Controlled Storage
    for i in range(len(chromo)):
        if i in temp_ctrl_products and chromo[i] not in (4, 8):
            fitness += 9

    # 5.5 - Fast-Moving Product Accessibility
    for i in range(len(chromo)):
        if chromo[i] != 6 and i in high_demand_products:
            fitness += 5

    # 5.6 - Heavy Item Placement
    for i in range(len(chromo)):
        if chromo[i] != 1 and i in heavy_products:
            fitness += 4

    # 5.7 - Product Compatibility
    categories = {}
    for i in range(len(chromo)):
        cat = products[i]["category"]
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(i)
    for cat, indices in categories.items():
        for a, b in combinations(indices, 2):
            if chromo[a] != chromo[b]:
                fitness += 3

    # 5.8 - Food and Chemical Incompatibility
    for z in range(1, num_zones + 1):
        food_count = 0
        chem_count = 0
        
        for i in range(len(chromo)):
            if chromo[i] == z:           # product i is in zone z
                if i in food_products:
                    food_count += 1
                if i in chemical_products:
                    chem_count += 1
        
        fitness += 15 * food_count * chem_count

    # 5.9 - Refrigerated Loading Dock Restriction
    for i in range(len(chromo)):
        if chromo[i] == 8 and i not in z8_eligible_products:
            fitness += 12


    return fitness

def calc_population_fitness(population):
    fitnesses = []
    for chromo in population:
        fit_val = calculate_fitness(chromo)
        fitnesses.append(fit_val)
    return fitnesses

def initialize_population(pop_size):
    population = [[random.randint(1, num_zones) for i in range(num_products)] for j in range(pop_size)]

    return population

def selection(population, fitnesses, k = 3):
    # Tournament size (k) is 3 by default
    parent1 = tournament_selection(population, fitnesses, k)
    parent2 = tournament_selection(population, fitnesses, k)
    
    return parent1, parent2

def tournament_selection(population, fitnesses, k):
    rand_indices = []
    combined_arr = []
    pop_size = len(population)

    while len(rand_indices) < k:
        ind = random.randint(0, pop_size - 1)
        if ind not in rand_indices:
            rand_indices.append(ind)
            combined_arr.append((population[ind], fitnesses[ind]))

    
    combined_arr.sort(key=lambda x: x[1])   # Sort by fitness (ascending)
    selected_chromo = combined_arr[0][0]    # Chromosome with lowest fitness

    return selected_chromo

def crossover(parent1, parent2):
    pt1 = random.randint(3, (num_products // 2) - (num_products // 4))                  # Crossover point 1
    pt2 = random.randint((num_products // 2) + (num_products // 4), num_products - 3)   # Crossover point 2

    child1 = parent1[:pt1] + parent2[pt1:pt2] + parent1[pt2:]
    child2 = parent2[:pt1] + parent1[pt1:pt2] + parent2[pt2:]
    
    return child1, child2

def mutation(chromosome, mutation_rate):
    for i in range(len(chromosome)):
        if random.random() < mutation_rate:     # Returns between 0 and 1
            chromosome[i] = random.randint(1, num_zones)
    return chromosome

# BONUS TASK: Excel Output
def export_to_excel(best_chromo, best_fitness, filename="warehouse_storage_plan.xlsx"):
    
    # ─────────────────── Sheet 1: Product Allocation ───────────────────
    rows = []
    for i in range(num_products):
        zone = best_chromo[i]
        p = products[i]
        rows.append({
            "Product":          p["name"],
            "Weight (kg)":      p["weight"],
            "Category":         p["category"],
            "Assigned Zone":    f"Z{zone}",
            "Zone Name":        zones[zone]["name"],
            "Fragile":          "Yes" if p["fragile"] else "No",
            "Hazardous":        "Yes" if p["hazardous"] else "No",
            "Temp Controlled":  "Yes" if p["temp_controlled"] else "No",
            "High Demand":      "Yes" if p["high_demand"] else "No",
        })
    df1 = pd.DataFrame(rows)

    # ────────────────────── Sheet 2: Zone Summary ──────────────────────
    zone_rows = []
    for z in range(1, num_zones + 1):
        zone_prods  = [products[i]["name"] for i in range(num_products) if best_chromo[i] == z]
        used_weight = sum(products[i]["weight"] for i in range(num_products) if best_chromo[i] == z)
        capacity    = zones[z]["capacity"]
        zone_rows.append({
            "Zone":             f"Z{z}",
            "Zone Name":        zones[z]["name"],
            "Capacity (kg)":    capacity,
            "Used (kg)":        used_weight,
            "Remaining (kg)":   capacity - used_weight,
            "Products Stored":  ", ".join(zone_prods) if zone_prods else "(empty)",
        })
    df2 = pd.DataFrame(zone_rows)

    # ──────────────────────── Sheet 3: Summary ─────────────────────────
    df3 = pd.DataFrame([
        {"Metric": "Best Fitness Score",  "Value": best_fitness},
        {"Metric": "Total Products",      "Value": num_products},
        {"Metric": "Total Zones",         "Value": num_zones},
        {"Metric": "Total Weight (kg)",   "Value": sum(p["weight"] for p in products)},
    ])

    # ─────────────── Write all sheets to one Excel file ────────────────
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df1.to_excel(writer, sheet_name="Product Allocation", index=False)
        df2.to_excel(writer, sheet_name="Zone Summary",       index=False)
        df3.to_excel(writer, sheet_name="Summary",            index=False)

    # ──────────────── Applying formatting ──────────────────────────────
    wb = load_workbook(filename)

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="2E75B6")
    normal_font  = Font(name="Arial", size=10)
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin         = Side(style="thin", color="CCCCCC")
    all_borders  = Border(left=thin, right=thin, top=thin, bottom=thin)
    green_fill   = PatternFill("solid", start_color="C6EFCE")  # Yes
    red_fill     = PatternFill("solid", start_color="FFC7CE")  # over capacity
    alt_fill_1   = PatternFill("solid", start_color="F2F7FD")  # alternating row
    alt_fill_2   = PatternFill("solid", start_color="FFFFFF")

    def format_sheet(ws, col_widths):
        # format header row
        for col, width in enumerate(col_widths, start=1):
            cell = ws.cell(row=1, column=col)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            cell.border    = all_borders
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        ws.row_dimensions[1].height = 20

        # format data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            row_fill = alt_fill_1 if row[0].row % 2 == 0 else alt_fill_2
            for cell in row:
                cell.font      = normal_font
                cell.alignment = center
                cell.border    = all_borders
                cell.fill      = row_fill

    # ─────────────────── Sheet 1: Product Allocation ───────────────────
    ws1 = wb["Product Allocation"]
    format_sheet(ws1, [22, 13, 13, 10, 28, 10, 12, 17, 13])

    # green highlight for Yes values in columns 6-9
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=6, max_col=9):
        for cell in row:
            if cell.value == "Yes":
                cell.fill = green_fill

    # ────────────────────── Sheet 2: Zone Summary ──────────────────────
    ws2 = wb["Zone Summary"]
    format_sheet(ws2, [8, 28, 14, 12, 15, 48])

    # red row if zone is over capacity
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        used      = row[3].value   # "Used (kg)" column
        capacity  = row[2].value   # "Capacity (kg)" column
        if used is not None and capacity is not None and used > capacity:
            for cell in row:
                cell.fill = red_fill
            row[4].font = Font(name="Arial", size=10, bold=True, color="FF0000")  # Remaining in red

    # ──────────────────────── Sheet 3: Summary ─────────────────────────
    ws3 = wb["Summary"]
    format_sheet(ws3, [25, 15])

    wb.save(filename)
    print(f"Excel report saved: {filename}")

if __name__ == "__main__":
    # Initialization
    pop_size = 200
    max_gen = 400
    tournament_size = 3
    base_mutation_rate = 0.05   # 5% chance for each gene to mutate
    mutation_rate = base_mutation_rate
    num_pairs_per_gen = 20      # 20 pairs = 40 children per generation
    no_improvement_count = 0
    no_improvement_counter = 0
    max_count = 200

    population = initialize_population(pop_size)

    print("Initial Population: ")
    for i in range(min(len(population), 5)):
        print(population[i])
    print("...\n")

    best_chromo = None
    best_fitness = float('inf')
    
    for gen in range(max_gen):
        # Calculate all fitnesses
        fitnesses = calc_population_fitness(population)
        
        # Keep track of best fitness so far
        gen_best_fitness = min(fitnesses)
        elite = population[fitnesses.index(gen_best_fitness)]        # Keeping best chromosome (elitism)
        if gen_best_fitness < best_fitness:
            best_fitness = gen_best_fitness
            best_chromo = population[fitnesses.index(gen_best_fitness)]
            no_improvement_count = 0
            no_improvement_counter = 0
            mutation_rate = base_mutation_rate
            print(f"Gen {gen + 1}: New best fitness = {best_fitness}")
        else:
            no_improvement_count += 1
            no_improvement_counter += 1
            if no_improvement_count > 50:
                mutation_rate = min(0.4, mutation_rate + 0.05)  # Increase mutation rate if no improvement for 50 generations (up to 40%)
                no_improvement_count = 0                        # Reset counter after increasing mutation rate
            if (gen + 1) % 100 == 0:
                print(f"Gen {gen + 1}: No improvement for {no_improvement_counter} generations, mutation rate = {mutation_rate:.2f}")

        # Check if goal reached and stop if so
        if best_fitness == 0:
            print(f"Best solution found at generation {gen + 1}!")
            break

        # Selection, Crossover, Mutation
        for skjfk in range(num_pairs_per_gen):
            parent1, parent2 = selection(population, fitnesses, tournament_size)
            child1, child2 = crossover(parent1, parent2)
            child1 = mutation(child1, mutation_rate)
            child2 = mutation(child2, mutation_rate)
            population.append(child1)
            population.append(child2)
            fitnesses = calc_population_fitness(population)
        
        # Add elite chromosome back to population
        population.append(elite)   

        # Sort and keep best chromosomes, discard the bad ones
        population.sort(key=lambda chromo: calculate_fitness(chromo))   # Sort population by fitness (ascending)
        population = population[:max_count]                             # Keep only the top 'max_count' chromosomes
        
    best_chromo = population[0]
    best_fitness = calculate_fitness(best_chromo)

    # Best Solution Possible: 78
    print('\n===== FINAL RESULTS =====')
    print(f"Generations: {gen + 1}")
    print(f'Best Chromosome: {best_chromo}')
    print(f'Best Fitness: {best_fitness}')

    print('\n===== OPTIMAL STORAGE PLAN =====')
    for z in range(1, num_zones + 1):
        zone_products = [products[i]["name"] for i in range(num_products) if best_chromo[i] == z]
        if zone_products:
            print(f'Z{z}: {", ".join(zone_products)}')
        else:
            print(f'Z{z}: (empty)')

    # Export results to Excel
    export_to_excel(best_chromo, best_fitness)

    pass